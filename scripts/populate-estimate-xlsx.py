#!/usr/bin/env python3
"""
Populate the QED42 Estimation Template from an optimistic estimate markdown file.

Usage:
    python3 scripts/populate-estimate-xlsx.py estimates/optimistic-estimate.md

The script parses the markdown tables in the estimate file and writes rows
to the Backend, Frontend, Fixed Cost Items, and AI tabs of the Excel template.

Tab detection: The script looks for '# Backend Tab', '# Frontend Tab',
'# Fixed Cost Items Tab', and '# AI Tab' section headers in the markdown.

Column mappings per tab:
    Backend:          B=Task, C=Description, E=Hours, F=Conf, G=Low Hrs, H=High Hrs, I=Assumptions, J=Proposed Solution, K=Reference Links
    Frontend:         B=Task, C=Description, E=Hours, F=Conf, G=Low Hrs, H=High Hrs, I=Assumptions, J=Exclusions, K=Reference Links
    Fixed Cost Items: B=Task, C=Description, E=Hours, F=Conf, G=Low Hrs, H=High Hrs, K=Assumptions, M=Reference Links
    AI:               (same as Backend if sheet exists)

Row 6 = header row, data starts at row 7 in all tabs.
"""

import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# Column mappings per tab (1-indexed column numbers)
TAB_COLUMNS = {
    'Backend': {
        'task': 2,        # B
        'description': 3, # C
        'hours': 5,       # E
        'conf': 6,        # F
        'low_hrs': 7,     # G
        'high_hrs': 8,    # H
        'assumptions': 9, # I
        'solution': 10,   # J (Proposed Solution)
        'links': 11,      # K
    },
    'Frontend': {
        'task': 2,        # B
        'description': 3, # C
        'hours': 5,       # E
        'conf': 6,        # F
        'low_hrs': 7,     # G
        'high_hrs': 8,    # H
        'assumptions': 9, # I
        'exclusions': 10, # J (Exclusions)
        'links': 11,      # K
    },
    'Fixed Cost Items': {
        'task': 2,         # B
        'description': 3,  # C
        'hours': 5,        # E
        'conf': 6,         # F
        'low_hrs': 7,      # G
        'high_hrs': 8,     # H
        'assumptions': 11, # K
        'links': 13,       # M
    },
}


def parse_estimate_markdown(md_path: str) -> dict[str, list[dict]]:
    """Parse estimate markdown into tab-separated row lists."""
    content = Path(md_path).read_text()

    # Split into tab sections using '# Tab Name' headers
    tab_sections = {}
    tab_pattern = re.compile(r'^# (Backend|Frontend|Fixed Cost Items|AI) Tab', re.MULTILINE)
    matches = list(tab_pattern.finditer(content))

    for i, match in enumerate(matches):
        tab_name = match.group(1)
        start = match.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(content)
        tab_sections[tab_name] = content[start:end]

    result = {}
    for tab_name, section_content in tab_sections.items():
        result[tab_name] = _parse_tab_section(tab_name, section_content)

    return result


def _parse_tab_section(tab_name: str, content: str) -> list[dict]:
    """Parse a single tab section into rows."""
    rows = []
    current_domain = None

    # Split by ## sub-headers within the tab
    sections = re.split(r'^## ', content, flags=re.MULTILINE)

    # Skip non-estimate sections
    skip_sections = {
        'Estimation Approach', 'Summary', 'Assumption Register',
        'Questions Resolved via Assumptions', 'Confidence Assessment',
        'Coverage Checklist',
    }

    for section in sections:
        if not section.strip():
            continue

        lines = section.strip().split('\n')
        title = lines[0].strip()

        if title in skip_sections:
            continue

        current_domain = title

        # Determine number of expected columns based on tab
        # Backend/AI: Task | Description | Hours | Conf | Low Hrs | High Hrs | Assumptions | Proposed Solution | Reference Links (9 cols)
        # Frontend: Task | Description | Hours | Conf | Low Hrs | High Hrs | Assumptions | Exclusions | Reference Links (9 cols)
        # Fixed Cost: Task | Description | Hours | Conf | Low Hrs | High Hrs | Assumptions | Reference Links (8 cols)

        if tab_name == 'Fixed Cost Items':
            row_pattern = re.compile(
                r'^\|([^|]+)\|([^|]+)\|([^|]+)\|([^|]+)\|([^|]+)\|([^|]+)\|([^|]+)\|([^|]*)\|$',
                re.MULTILINE,
            )
        else:
            row_pattern = re.compile(
                r'^\|([^|]+)\|([^|]+)\|([^|]+)\|([^|]+)\|([^|]+)\|([^|]+)\|([^|]+)\|([^|]+)\|([^|]*)\|$',
                re.MULTILINE,
            )

        for match in row_pattern.finditer(section):
            task = match.group(1).strip()
            description = match.group(2).strip()
            hours_str = match.group(3).strip()
            conf_str = match.group(4).strip()

            # Skip template placeholder rows, header rows, and separator rows
            if task.startswith('[') or task == 'Task' or task.startswith('---') or task.startswith('**'):
                continue

            # Parse hours
            try:
                hours = float(hours_str)
            except ValueError:
                hours = 0

            # Parse conf
            try:
                conf = float(conf_str)
            except ValueError:
                conf = 0

            if hours == 0 and not description:
                continue

            # Parse Low Hrs and High Hrs (groups 5 and 6 in all tabs)
            low_hrs_str = match.group(5).strip()
            high_hrs_str = match.group(6).strip()
            try:
                low_hrs = float(low_hrs_str)
            except ValueError:
                low_hrs = hours
            try:
                high_hrs = float(high_hrs_str)
            except ValueError:
                high_hrs = hours

            if tab_name == 'Fixed Cost Items':
                assumptions = match.group(7).strip()
                links = match.group(8).strip()
                rows.append({
                    'domain': current_domain,
                    'task': task,
                    'description': description,
                    'hours': hours,
                    'conf': conf,
                    'low_hrs': low_hrs,
                    'high_hrs': high_hrs,
                    'assumptions': assumptions.replace('<br>', '\n'),
                    'links': links,
                })
            else:
                assumptions = match.group(7).strip()
                col6 = match.group(8).strip()  # Proposed Solution or Exclusions
                links = match.group(9).strip()
                rows.append({
                    'domain': current_domain,
                    'task': task,
                    'description': description,
                    'hours': hours,
                    'conf': conf,
                    'low_hrs': low_hrs,
                    'high_hrs': high_hrs,
                    'assumptions': assumptions.replace('<br>', '\n'),
                    'col6': col6,  # Proposed Solution (Backend/AI) or Exclusions (Frontend)
                    'links': links,
                })

    return rows


def populate_tab(wb, tab_name: str, rows: list[dict]) -> None:
    """Write rows to a specific tab in the workbook."""
    # Map AI tab to Backend column layout
    sheet_name = tab_name
    col_map = TAB_COLUMNS.get(tab_name, TAB_COLUMNS.get('Backend'))

    if sheet_name not in wb.sheetnames:
        print(f"Warning: '{sheet_name}' sheet not found — skipping {len(rows)} rows")
        return

    ws = wb[sheet_name]
    start_row = 7
    current_domain = None
    row_idx = start_row

    for item in rows:
        # Insert domain header row if domain changed
        if item['domain'] != current_domain:
            current_domain = item['domain']
            ws.cell(row=row_idx, column=col_map['task'], value=current_domain)
            row_idx += 1

        # Write common columns
        ws.cell(row=row_idx, column=col_map['task'], value=item['task'])
        ws.cell(row=row_idx, column=col_map['description'], value=item['description'])
        ws.cell(row=row_idx, column=col_map['hours'], value=item['hours'])
        ws.cell(row=row_idx, column=col_map['conf'], value=item['conf'])
        ws.cell(row=row_idx, column=col_map['low_hrs'], value=item.get('low_hrs', item['hours']))
        ws.cell(row=row_idx, column=col_map['high_hrs'], value=item.get('high_hrs', item['hours']))
        ws.cell(row=row_idx, column=col_map['assumptions'], value=item['assumptions'])

        # Tab-specific columns
        if tab_name == 'Fixed Cost Items':
            ws.cell(row=row_idx, column=col_map['links'], value=item.get('links', ''))
        else:
            col6_key = 'solution' if tab_name in ('Backend', 'AI') else 'exclusions'
            if col6_key in col_map:
                ws.cell(row=row_idx, column=col_map[col6_key], value=item.get('col6', ''))
            ws.cell(row=row_idx, column=col_map['links'], value=item.get('links', ''))

        row_idx += 1

    low_total = sum(r.get('low_hrs', r['hours']) for r in rows)
    high_total = sum(r.get('high_hrs', r['hours']) for r in rows)
    avg_total = (low_total + high_total) / 2 if rows else 0
    print(f"  {tab_name}: {len(rows)} line items, Low={low_total}h / High={high_total}h / Avg={avg_total}h")


def populate_xlsx(tab_data: dict[str, list[dict]], xlsx_path: str) -> str:
    """Write parsed rows to all tabs of the Excel template."""
    wb = openpyxl.load_workbook(xlsx_path)

    for tab_name, rows in tab_data.items():
        if rows:
            populate_tab(wb, tab_name, rows)

    output_path = xlsx_path.replace('.xlsx', '-optimistic.xlsx')
    wb.save(output_path)
    return output_path


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 scripts/populate-estimate-xlsx.py <estimate-markdown-file>")
        print("Example: python3 scripts/populate-estimate-xlsx.py estimates/optimistic-estimate.md")
        sys.exit(1)

    md_path = sys.argv[1]
    if not Path(md_path).exists():
        print(f"Error: File not found: {md_path}")
        sys.exit(1)

    # Find the Excel template
    xlsx_path = 'estimation_template/QED42-Estimate_Template.xlsx'
    if not Path(xlsx_path).exists():
        from glob import glob
        candidates = glob('estimation_template/*.xlsx')
        if candidates:
            xlsx_path = candidates[0]
        else:
            print("Error: No Excel template found in estimation_template/")
            sys.exit(1)

    print(f"Parsing estimate: {md_path}")
    tab_data = parse_estimate_markdown(md_path)

    total_rows = sum(len(rows) for rows in tab_data.values())
    print(f"Found {total_rows} estimate line items across {len(tab_data)} tabs")

    if total_rows == 0:
        print("Warning: No estimate rows found. Check markdown format.")
        sys.exit(1)

    print(f"Populating Excel template: {xlsx_path}")
    output_path = populate_xlsx(tab_data, xlsx_path)
    print(f"\nOutput saved to: {output_path}")

    total_low = sum(
        sum(r.get('low_hrs', r['hours']) for r in rows)
        for rows in tab_data.values()
    )
    total_high = sum(
        sum(r.get('high_hrs', r['hours']) for r in rows)
        for rows in tab_data.values()
    )
    total_avg = (total_low + total_high) / 2
    print(f"Total: Low={total_low}h / High={total_high}h / Avg={total_avg}h")


if __name__ == '__main__':
    main()
